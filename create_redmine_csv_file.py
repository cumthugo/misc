import sys
reload(sys)
sys.setdefaultencoding('utf8')
import csv
import re
from openpyxl import load_workbook


wb = load_workbook(filename=r'all.xlsx')
ws = wb.get_sheet_by_name('all')

rows = ws.rows
cols = ws.columns

priority_str = {'High':'高','Medium': '普通', 'Low':'低','Unassigned':'普通','Unclassified':'普通'}
serverity_str = {'High':'P1','Medium':'P2','Low':'P3','Unassigned':'P2','Unclassified':'P2'}
version_str = {u'1':'328_YF',u'2':'K211_YF',u'3':'K216',u'4':'K256',u'5':'K215'}

w_file = open('csv_for_redmine_import.csv','wb') 
csv_writer = csv.writer(w_file)
Header_str = ['#','项目','跟踪','状态','优先级','主题','优先级(HMI12)','发现版本','提出人员','修正版本','SVN Revision','对策确认版本','责任方(HMI12)','关联No','Car(HMI12)','私有','描述']
csv_writer.writerow(Header_str)
#we need to write every row value
for i in range(2,len(rows)+1):
    v_priority = ws.cell('K%d'%i).value
    summary = ws.cell('B%d'%i).value
    serverity = ws.cell('C%d'%i).value
    description = ws.cell('D%d'%i).value
    comment = ws.cell('E%d'%i).value
    foundin = ws.cell('I%d'%i).value
    defectid = ws.cell('A%d'%i).value
    tags = ws.cell('H%d'%i).value

    if defectid == 556032:
        print "defectid = ", defectid
        version = version_str[u'5'] #default is K215
        version_ret = re.findall('[0-9]\.[0-9]\.0([1-5])',description)
        if version_ret:
            version = version_str[version_ret[0][0]]
        item = ['48144','HMI12','ST Bug','新建',priority_str[v_priority],summary,serverity_str[serverity],foundin,'YFVE_Z 张勇','Ver*.***','rev.0000; rev.0000','Ver*.***','HMI',defectid,version,'否','%s\n\nComments:\n%s'%(description,comment)] # TODO: 'HMI' and 'NAF' need to judge
        csv_writer.writerow(item)
w_file.close()
print 'done'
